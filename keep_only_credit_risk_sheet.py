from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("data/credit-risk-only.xlsx")
TARGET_SHEET = "Credit risk"


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH)
    if TARGET_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{TARGET_SHEET}' not found. Available sheets: {workbook.sheetnames}"
        )

    for sheet_name in list(workbook.sheetnames):
        if sheet_name != TARGET_SHEET:
            del workbook[sheet_name]

    workbook.active = 0
    workbook.save(WORKBOOK_PATH)
    print(f"Saved workbook with only '{TARGET_SHEET}' sheet: {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
